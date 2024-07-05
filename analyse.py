import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

def highlight_differences(file_path, sheet_name, column, values_to_highlight, color):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    for row in ws[column]:
        if row.value in values_to_highlight:
            row.fill = fill
    
    wb.save(file_path)

# Lire les fichiers Excel
df_output = pd.read_excel('C3_accredited_users.xlsx')
df_jerome = pd.read_excel('collegue.xlsx')

# Extraire les colonnes à comparer
col_output = df_output['GROUP MAIL'].dropna().str.lower()
col_jerome = df_jerome['B'].dropna().str.lower()

# Trouver les différences et les éléments communs
set_output = set(col_output)
set_jerome = set(col_jerome)

common = set_output & set_jerome
only_in_output = set_output - set_jerome
only_in_jerome = set_jerome - set_output

# Créer un DataFrame pour l'analyse détaillée
analysis = pd.DataFrame({
    'Utilisateur': list(set_output | set_jerome),
    'Dans output_cleaned': [user in set_output for user in (set_output | set_jerome)],
    'Dans jerome': [user in set_jerome for user in (set_output | set_jerome)]
})

# Trier le DataFrame
analysis = analysis.sort_values(['Dans output_cleaned', 'Dans jerome', 'Utilisateur'])

# Sauvegarder l'analyse dans un nouveau fichier Excel
analysis.to_excel('analyse_détaillée.xlsx', index=False)

# Mettre en évidence les différences dans les fichiers originaux
highlight_differences('output_cleaned.xlsx', 'Sheet1', 'A', only_in_output, 'FFFF00')  # Jaune
highlight_differences('jerome.xlsx', 'Sheet1', 'B', only_in_jerome, 'FF9900')  # Orange

# Afficher les résultats
print(f"Analyse détaillée sauvegardée dans 'analyse_détaillée.xlsx'")
print(f"\nUtilisateurs communs: {len(common)}")
print(f"Utilisateurs uniquement dans output_cleaned: {len(only_in_output)}")
print(f"Utilisateurs uniquement dans jerome: {len(only_in_jerome)}")

# Calculer les statistiques
total_output = len(col_output)
total_jerome = len(col_jerome)

print(f"\nStatistiques:")
print(f"Total utilisateurs dans output_cleaned: {total_output}")
print(f"Total utilisateurs dans jerome: {total_jerome}")
print(f"Pourcentage de correspondance: {(len(common) / max(total_output, total_jerome)) * 100:.2f}%")

# Afficher quelques exemples de différences
print("\nExemples d'utilisateurs uniquement dans output_cleaned:")
print(list(only_in_output)[:5])
print("\nExemples d'utilisateurs uniquement dans jerome:")
print(list(only_in_jerome)[:5])