import pandas as pd
from alive_progress import alive_bar
from openpyxl import load_workbook

def read_excel_with_progress(file_path, sheet_name=None, header='infer'):
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    total_rows = ws.max_row
    wb.close()

    with alive_bar(total_rows, title=f"Chargement du fichier '{file_path}'") as bar:
        data = pd.read_excel(file_path, sheet_name=sheet_name, header=header)
        for _ in range(total_rows):
            bar()
    return data

def clean_department(dept):
    if isinstance(dept, str):
        return dept.split('.')[0].strip()
    return dept

def is_excluded(row, excluded_departments, excluded_emails, all_types_departments):
    """Vérifie si une ligne doit être exclue selon les nouvelles règles."""
    excluded_contract_codes = {'999', 'VIE', 'STA', 'APP'}
    excluded_labels = {
        'Trainee', 'Apprenticeship/learning', 'Extern-temp', 'International Business Volunte'
    }
    # Ne pas exclure si le département est dans all_types_departments
    if row['LIB_SERVICE'] in all_types_departments or row['LIB_CENTRE_ACTIVITE'] in all_types_departments:
        return False
    return (
        row['CONTRACT_GROUP_TYPE_CODE'] in excluded_contract_codes or
        any(label in str(row['CONTRACT_GROUP_TYPE_LABEL']) for label in excluded_labels) or
        '-ext' in str(row['GROUP_MAIL']) or
        row['LIB_SERVICE'] in excluded_departments or
        row['GROUP_MAIL'] in excluded_emails or
        str(row['STATUS_GROUP_LABEL']) == 'Absents'
    )

print("Initialisation du script...")

# Lire les fichiers Excel avec barres de progression
print("\n---------------\nLecture des fichiers Excel...\n---------------")
people = read_excel_with_progress('people.xlsx', sheet_name=0, header=0)
custom = read_excel_with_progress('custom.xlsx', sheet_name=0, header=0)
departements_c3 = read_excel_with_progress('departements.xlsx', sheet_name='LIST C3 DPT ONLY INTERNALS', header=None)
nominative_users = read_excel_with_progress('departements.xlsx', sheet_name='NOMINATIVE USERS', header=0)
elr_habilite = read_excel_with_progress('departements.xlsx', sheet_name='LIST OF ELR', header=0)
dpts_user_to_exclude = read_excel_with_progress('departements.xlsx', sheet_name='DPTS-USER TO BE EXCLUDED', header=0)
all_types_departments = read_excel_with_progress('departements.xlsx', sheet_name='LIST C3 DPT ALL TYPES', header=None)

print("\n---------------\nVérification des colonnes dans les fichiers Excel...\n---------------")
print("Colonnes de 'people':", people.columns.tolist())
print("Colonnes de 'custom':", custom.columns.tolist())
print("Colonnes de 'departements_c3':", departements_c3.columns.tolist())
print("Colonnes de 'NOMINATIVE USERS':", nominative_users.columns.tolist())
print("Colonnes de 'LIST OF ELR':", elr_habilite.columns.tolist())
print("Colonnes de 'DPTS-USER TO BE EXCLUDED':", dpts_user_to_exclude.columns.tolist())
print("Colonnes de 'LIST C3 DPT ALL TYPES':", all_types_departments.columns.tolist())

print("\n---------------\nAperçu des premières lignes des DataFrames...\n---------------")
print("Premières lignes de 'people':\n", people.head())
print("Premières lignes de 'custom':\n", custom.head())
print("Premières lignes de 'departements_c3':\n", departements_c3.head())
print("Premières lignes de 'NOMINATIVE USERS':\n", nominative_users.head())
print("Premières lignes de 'LIST OF ELR':\n", elr_habilite.head())
print("Premières lignes de 'DPTS-USER TO BE EXCLUDED':\n", dpts_user_to_exclude.head())
print("Premières lignes de 'LIST C3 DPT ALL TYPES':\n", all_types_departments.head())

# Étape 1 : Synchronisation des données entre people et custom
print("\n---------------\nSynchronisation des données entre 'people' et 'custom'...\n---------------")
custom_copy = custom.copy()
custom_copy.rename(columns={'GGI': 'IGG', 'Email': 'GROUP_MAIL', 'Department': 'LIB_SERVICE'}, inplace=True)

merged_data = pd.merge(people, custom_copy[['IGG', 'GROUP_MAIL', 'LIB_SERVICE']], on='IGG', how='left', suffixes=('', '_custom'))
merged_data['GROUP_MAIL'] = merged_data['GROUP_MAIL'].fillna(merged_data['GROUP_MAIL_custom'])
merged_data['LIB_SERVICE'] = merged_data['LIB_SERVICE'].fillna(merged_data['LIB_SERVICE_custom'])
merged_data.drop(columns=['GROUP_MAIL_custom', 'LIB_SERVICE_custom'], inplace=True)

# Enregistrer le fichier synchronisé
merged_data.to_excel('synchronized_people_custom.xlsx', index=False)
print("\n---------------\nLe fichier 'synchronized_people_custom.xlsx' a été créé avec succès.\n---------------")

# Récupérer les départements et emails à exclure
excluded_departments = set(dpts_user_to_exclude.iloc[:, 0].dropna().apply(clean_department))
excluded_emails = set(dpts_user_to_exclude.iloc[:, 4].dropna())
all_types_departments_set = set(all_types_departments.iloc[:, 0].dropna().apply(clean_department))
print(f"\nDépartements à exclure : {excluded_departments}")
print(f"\nEmails à exclure : {excluded_emails}")
print(f"\nDépartements 'ALL TYPES' : {all_types_departments_set}")

# Étape 2 : Filtrage selon 'LIST C3 DPT ONLY INTERNALS'
print("\n---------------\nFiltrage selon 'LIST C3 DPT ONLY INTERNALS'...\n---------------")
departements_c3_clean = departements_c3.iloc[5:, 0].apply(clean_department)
c3_departments = set(departements_c3_clean)
print(f"\nDépartements C3 nettoyés : {c3_departments}")

def get_c3_department(row):
    lib_service = str(row['LIB_SERVICE']) if pd.notna(row['LIB_SERVICE']) else ''
    lib_centre_activite = str(row['LIB_CENTRE_ACTIVITE']) if pd.notna(row['LIB_CENTRE_ACTIVITE']) else ''
    for dept in c3_departments:
        if isinstance(dept, str) and (lib_service.startswith(dept) or lib_centre_activite.startswith(dept)):
            return dept
    return None

merged_data['DEPARTEMENT'] = merged_data.apply(get_c3_department, axis=1)
filtered_data_c3 = merged_data[merged_data['DEPARTEMENT'].notna()]
print("\n---------------\nAperçu des données filtrées (départements C3)...\n---------------")
print("Premières lignes de 'filtered_data_c3':\n", filtered_data_c3.head())

# Appliquer les nouvelles règles d'exclusion
filtered_data_c3 = filtered_data_c3[~filtered_data_c3.apply(lambda row: is_excluded(row, excluded_departments, excluded_emails, all_types_departments_set), axis=1)]
print("\n---------------\nAperçu des données filtrées après exclusion (départements C3)...\n---------------")
print("Premières lignes de 'filtered_data_c3' après exclusion:\n", filtered_data_c3.head())

# Étape 3 : Filtrage selon 'LIST OF ELR'
print("\n---------------\nFiltrage selon 'LIST OF ELR'...\n---------------")
elr_habilite_c3 = set(elr_habilite['ELR habilité au C3'])
filtered_data_elr = merged_data[merged_data['LIB_ELR_RAPPRO'].isin(elr_habilite_c3)]
print("\n---------------\nAperçu des données filtrées (ELR habilité au C3)...\n---------------")
print("Premières lignes de 'filtered_data_elr':\n", filtered_data_elr.head())

# Appliquer les nouvelles règles d'exclusion
filtered_data_elr = filtered_data_elr[~filtered_data_elr.apply(lambda row: is_excluded(row, excluded_departments, excluded_emails, all_types_departments_set), axis=1)]
print("\n---------------\nAperçu des données filtrées après exclusion (ELR habilité au C3)...\n---------------")
print("Premières lignes de 'filtered_data_elr' après exclusion:\n", filtered_data_elr.head())

# Étape 4 : Application des autres filtres (entités à partir de la colonne G)
print("\n---------------\nApplication des autres filtres (emails nominatives et entités)...\n---------------")
nominative_emails = set(nominative_users['Mail'])
nominative_entites = set(nominative_users.iloc[:, 6])  # Colonne G correspond à l'index 6
filtered_data_others = merged_data[merged_data['GROUP_MAIL'].isin(nominative_emails) | merged_data['LIB_SERVICE'].isin(nominative_entites)]
print("\n---------------\nAperçu des données filtrées (emails nominatives et entités)...\n---------------")
print("Premières lignes de 'filtered_data_others':\n", filtered_data_others.head())

# Combiner tous les filtres pour inclure les utilisateurs qui remplissent au moins un critère
filtered_data = pd.concat([filtered_data_c3, filtered_data_elr, filtered_data_others]).drop_duplicates()

# Filtre final : Exclusion des utilisateurs avec 'Absents' dans STATUS_GROUP_LABEL et ceux dans DPTS-USER TO BE EXCLUDED
filtered_data = filtered_data[filtered_data['STATUS_GROUP_LABEL'] != 'Absents']  # Exclure les utilisateurs avec STATUS_GROUP_LABEL == 'Absents'
filtered_data = filtered_data[~filtered_data.apply(lambda row: is_excluded(row, excluded_departments, excluded_emails, all_types_departments_set), axis=1)]

print("\n---------------\nAperçu des données combinées après filtres...\n---------------")
print("Premières lignes de 'filtered_data':\n", filtered_data.head())

# Sélectionner les colonnes nécessaires pour le fichier final
final_data = filtered_data[['IGG', 'GROUP_MAIL', 'DEPARTEMENT']]

# Sauvegarder le fichier final
final_data.to_excel('C3_accredited_users.xlsx', index=False)
print("\n---------------\nLe fichier 'C3_accredited_users.xlsx' a été créé avec succès. Le processus est terminé.\n---------------")
