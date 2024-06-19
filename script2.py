import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook

def read_excel_with_progress(file_path, header=True):
    # Charger le workbook avec openpyxl
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active  # Charger la première feuille active

    if header:
        # Lire les noms des colonnes depuis la première ligne
        columns = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        data_start_row = 2
    else:
        # Supposer des noms de colonne génériques si aucune entête n'est présente
        columns = [f'Column_{i+1}' for i in range(len(next(ws.iter_rows(min_row=1, max_row=1))))]
        data_start_row = 1

    # Initialiser un dictionnaire pour recueillir les données
    data = {column: [] for column in columns}

    # Lire les données avec progression
    row_count = ws.max_row - data_start_row + 1  # Calcul du nombre total de lignes à lire pour la progression
    for row in tqdm(ws.iter_rows(min_row=data_start_row, max_row=ws.max_row), total=row_count, desc=f"Lecture de {file_path}"):
        for key, cell in zip(columns, row):
            data[key].append(cell.value)

    # Fermer le workbook
    wb.close()

    # Créer un DataFrame à partir du dictionnaire
    return pd.DataFrame(data)

# Utilisation de la fonction pour lire les fichiers
people = read_excel_with_progress('people.xlsx')
custom = read_excel_with_progress('custom.xlsx')
departements_c3 = read_excel_with_progress('departements.xlsx', header=False)

# Renommer les colonnes dans 'custom' pour correspondre à celles de 'people'
custom.rename(columns={'GGI': 'IGG', 'Email': 'GROUP_MAIL', 'Department': 'LIB_SERVICE'}, inplace=True)

# Fusionner people avec custom pour compléter les informations manquantes
merged_data = pd.merge(people, custom[['IGG', 'GROUP_MAIL', 'LIB_SERVICE']], on='IGG', how='left', suffixes=('', '_custom'))

# Utiliser fillna pour combler les informations manquantes
merged_data['GROUP_MAIL'] = merged_data['GROUP_MAIL'].fillna(merged_data['GROUP_MAIL_custom'])
merged_data['LIB_SERVICE'] = merged_data['LIB_SERVICE'].fillna(merged_data['LIB_SERVICE_custom'])

# Supprimer les colonnes inutiles après la fusion
merged_data.drop(columns=['GROUP_MAIL_custom', 'LIB_SERVICE_custom'], inplace=True)

# Filtrer les départements C3
c3_departments = set(departements_c3['Column_1'])  # Assurez-vous que 'Column_1' est la colonne correcte pour les départements C3
filtered_data = merged_data[merged_data['LIB_SERVICE'].isin(c3_departments)]

# Sélectionner les colonnes nécessaires pour le fichier final
final_data = filtered_data[['IGG', 'GROUP_MAIL', 'LIB_SERVICE']]

# Sauvegarder le fichier final
final_data.to_excel('C3_accredited_users.xlsx', index=False)

print("Le fichier 'C3_accredited_users.xlsx' a été créé avec succès.")
