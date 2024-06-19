import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook

def read_excel_with_progress(file_path, header='infer'):
    # Charger le workbook avec openpyxl pour lire le nombre de lignes
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active
    total_rows = ws.max_row
    wb.close()

    # Afficher la barre de progression pendant le chargement
    with tqdm(total=total_rows, desc=f"Chargement du fichier '{file_path}'") as pbar:
        data = pd.read_excel(file_path, header=header)
        pbar.update(total_rows)
    return data

print("Initialisation du script...")

# Lire les fichiers Excel avec barres de progression
print("Lecture des fichiers Excel...")
people = read_excel_with_progress('people.xlsx', header=0)
custom = read_excel_with_progress('custom.xlsx', header=0)
departements_c3 = read_excel_with_progress('departements.xlsx', header=None)

# Vérification des colonnes des DataFrames
print("Colonnes de 'people':", people.columns.tolist())
print("Colonnes de 'custom':", custom.columns.tolist())

# Créer des copies des DataFrames pour les manipulations
print("Création de copies des DataFrames pour manipulation...")
people_copy = people.copy()
custom_copy = custom.copy()

# Renommer les colonnes dans 'custom_copy'
print("Renommage des colonnes dans 'custom' pour correspondre à celles de 'people'...")
if 'GGI' in custom_copy.columns and 'Email' in custom_copy.columns and 'Department' in custom_copy.columns:
    custom_copy.rename(columns={'GGI': 'IGG', 'Email': 'GROUP_MAIL', 'Department': 'LIB_SERVICE'}, inplace=True)
    print("Colonnes après renommage dans 'custom':", custom_copy.columns.tolist())
else:
    raise KeyError("Les colonnes attendues 'GGI', 'Email' et 'Department' ne sont pas présentes dans 'custom'.")

# Fusionner people_copy avec custom_copy pour compléter les informations manquantes
print("Fusion des DataFrames...")
merged_data = pd.merge(people_copy, custom_copy[['IGG', 'GROUP_MAIL', 'LIB_SERVICE']], on='IGG', how='left', suffixes=('', '_custom'))

# Utiliser fillna pour combler les informations manquantes
print("Comblement des informations manquantes...")
merged_data['GROUP_MAIL'] = merged_data['GROUP_MAIL'].fillna(merged_data['GROUP_MAIL_custom'])
merged_data['LIB_SERVICE'] = merged_data['LIB_SERVICE'].fillna(merged_data['LIB_SERVICE_custom'])

# Supprimer les colonnes inutiles après la fusion
print("Suppression des colonnes temporaires après la fusion...")
merged_data.drop(columns=['GROUP_MAIL_custom', 'LIB_SERVICE_custom'], inplace=True)

# Filtrer les départements C3
print("Filtration des utilisateurs selon les départements C3...")
c3_departments = set(departements_c3.iloc[:, 0])
filtered_data = merged_data[merged_data['LIB_SERVICE'].isin(c3_departments)].copy()

# Assurer que la colonne LIB_SERVICE est toujours remplie
print("Assurer que la colonne 'LIB_SERVICE' est toujours remplie...")
filtered_data.loc[:, 'LIB_SERVICE'] = filtered_data['LIB_SERVICE'].ffill()

# Sélectionner les colonnes nécessaires pour le fichier final
print("Sélection des colonnes finales pour l'export...")
final_data = filtered_data[['IGG', 'GROUP_MAIL', 'LIB_SERVICE']]

# Sauvegarder le fichier final
print("Sauvegarde du fichier final 'C3_accredited_users.xlsx'...")
final_data.to_excel('C3_accredited_users.xlsx', index=False)

print("Le fichier 'C3_accredited_users.xlsx' a été créé avec succès. Le processus est terminé.")
