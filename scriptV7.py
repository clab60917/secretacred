import pandas as pd
from alive_progress import alive_bar
from openpyxl import load_workbook

def read_excel_with_progress(file_path, sheet_name=None, header='infer'):
    # Charger le workbook avec openpyxl pour lire le nombre de lignes
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    total_rows = ws.max_row
    wb.close()

    # Afficher la barre de progression pendant le chargement
    with alive_bar(total_rows, title=f"Chargement du fichier '{file_path}'") as bar:
        data = pd.read_excel(file_path, sheet_name=sheet_name, header=header)
        for _ in range(total_rows):
            bar()
    return data

def clean_department(dept):
    """Nettoie les annotations des départements."""
    if isinstance(dept, str):
        return dept.split('.')[0].strip()
    return dept

print("Initialisation du script...")

# Lire les fichiers Excel avec barres de progression
print("\n---------------\nLecture des fichiers Excel...\n---------------")
people = read_excel_with_progress('people.xlsx', header=0)
custom = read_excel_with_progress('custom.xlsx', header=0)
departements_c3 = read_excel_with_progress('departements.xlsx', sheet_name='LIST C3 DPT ONLY INTERNALS', header=None)
nominative_users = read_excel_with_progress('departements.xlsx', sheet_name='NOMINATIVE USERS + ORIGINE', header=0)
elr_habilite = read_excel_with_progress('departements.xlsx', sheet_name='LIST OF ELR', header=0)

# Afficher les noms des colonnes pour vérification
print("\n---------------\nVérification des colonnes dans les fichiers Excel...\n---------------")
print("Colonnes de 'people':", people.columns.tolist())
print("Colonnes de 'custom':", custom.columns.tolist())
print("Colonnes de 'departements_c3':", departements_c3.columns.tolist())
print("Colonnes de 'NOMINATIVE USERS + ORIGINE':", nominative_users.columns.tolist())
print("Colonnes de 'LIST OF ELR':", elr_habilite.columns.tolist())

# Vérifier les premières lignes de chaque DataFrame
print("\n---------------\nAperçu des premières lignes des DataFrames...\n---------------")
print("Premières lignes de 'people':\n", people.head())
print("Premières lignes de 'custom':\n", custom.head())
print("Premières lignes de 'departements_c3':\n", departements_c3.head())
print("Premières lignes de 'NOMINATIVE USERS + ORIGINE':\n", nominative_users.head())
print("Premières lignes de 'LIST OF ELR':\n", elr_habilite.head())

# Nettoyer les départements C3 et créer un set des départements C3
departements_c3_clean = departements_c3.iloc[5:, 0].apply(clean_department)  # Commence à partir de la ligne 6
c3_departments = set(departements_c3_clean)

# Créer des sets pour les filtres supplémentaires
nominative_emails = set(nominative_users['Mail'])
elr_habilite_c3 = set(elr_habilite['ELR habilité au C3'])

# Créer des copies des DataFrames pour les manipulations
people_copy = people.copy()
custom_copy = custom.copy()

# Renommer les colonnes dans 'custom_copy'
custom_copy.rename(columns={'GGI': 'IGG', 'Email': 'GROUP_MAIL', 'Department': 'LIB_SERVICE'}, inplace=True)

# Fusionner people_copy avec custom_copy pour compléter les informations manquantes
merged_data = pd.merge(people_copy, custom_copy[['IGG', 'GROUP_MAIL', 'LIB_SERVICE']], on='IGG', how='left', suffixes=('', '_custom'))

# Utiliser fillna pour combler les informations manquantes
merged_data['GROUP_MAIL'] = merged_data['GROUP_MAIL'].fillna(merged_data['GROUP_MAIL_custom'])
merged_data['LIB_SERVICE'] = merged_data['LIB_SERVICE'].fillna(merged_data['LIB_SERVICE_custom'])

# Supprimer les colonnes inutiles après la fusion
merged_data.drop(columns=['GROUP_MAIL_custom', 'LIB_SERVICE_custom'], inplace=True)

# Nouveau filtre pour vérifier la colonne LIB_CENTRE_ACTIVITE si LIB_SERVICE ne contient pas de /
def get_c3_department(row):
    lib_service = str(row['LIB_SERVICE'])
    if '/' in lib_service:
        department = clean_department(lib_service)
        if department in c3_departments:
            return department
    else:
        department = clean_department(str(row['LIB_CENTRE_ACTIVITE']))
        if department in c3_departments:
            return department
    return None

# Appliquer le filtre et récupérer le département C3
merged_data['DEPARTEMENT'] = merged_data.apply(get_c3_department, axis=1)

# Filtrer les utilisateurs C3
filtered_data = merged_data[merged_data['DEPARTEMENT'].notna()]

# Appliquer les nouveaux filtres
filtered_data = filtered_data[filtered_data['GROUP_MAIL'].isin(nominative_emails) | filtered_data['LIB_ELR_RAPPRO'].isin(elr_habilite_c3)]

# Sélectionner les colonnes nécessaires pour le fichier final
final_data = filtered_data[['IGG', 'GROUP_MAIL', 'DEPARTEMENT']]

# Sauvegarder le fichier final
final_data.to_excel('C3_accredited_users.xlsx', index=False)

print("\n---------------\nLe fichier 'C3_accredited_users.xlsx' a été créé avec succès. Le processus est terminé.\n---------------")
