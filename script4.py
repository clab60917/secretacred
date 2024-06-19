import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook

def read_excel_with_progress(file_path, header='infer'):
    # Charger le workbook avec openpyxl pour lire le nombre de lignes
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active  # Charger la première feuille active

    # Préparation pour lire les noms des colonnes et les données
    if header is None:
        columns = [f'Column_{i+1}' for i in range(len(next(ws.iter_rows(min_row=1, max_row=1))))]
        data_start_row = 1
    else:
        columns = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        data_start_row = 2

    # Calculer le nombre total de lignes à lire pour la progression
    row_count = ws.max_row - data_start_row + 1

    # Initialiser un dictionnaire pour recueillir les données
    data = {column: [] for column in columns}

    # Lire les données avec progression
    pbar = tqdm(ws.iter_rows(min_row=data_start_row, max_row=ws.max_row), total=row_count, desc=f"Lecture de {file_path}")
    for row in pbar:
        for key, cell in zip(columns, row):
            data[key].append(cell.value)

    # Fermer le workbook
    wb.close()

    # Créer un DataFrame à partir du dictionnaire
    return pd.DataFrame(data, columns=columns)

print("Initialisation du script...")

# Lire les fichiers Excel avec barres de progression
people = read_excel_with_progress('people.xlsx')
custom = read_excel_with_progress('custom.xlsx')
departements_c3 = read_excel_with_progress('departements.xlsx', header=None)

# Créer des copies des DataFrames pour les manipulations
people_copy = people.copy()
custom_copy = custom.copy()

# Renommer les colonnes dans 'custom_copy' pour correspondre à celles de 'people_copy'
custom_copy.rename(columns={'GGI': 'IGG', 'Email': 'GROUP_MAIL', 'Department': 'LIB_SERVICE'}, inplace=True)

# Fusionner people_copy avec custom_copy pour compléter les informations manquantes
merged_data = pd.merge(people_copy, custom_copy[['IGG', 'GROUP_MAIL', 'LIB_SERVICE']], on='IGG', how='left', suffixes=('', '_custom'))

# Utiliser fillna pour combler les informations manquantes
merged_data['GROUP_MAIL'] = merged_data['GROUP_MAIL'].fillna(merged_data['GROUP_MAIL_custom'])
merged_data['LIB_SERVICE'] = merged_data['LIB_SERVICE'].fillna(merged_data['LIB_SERVICE_custom'])

# Supprimer les colonnes inutiles après la fusion
merged_data.drop(columns=['GROUP_MAIL_custom', 'LIB_SERVICE_custom'], inplace=True)

# Filtrer les départements C3
c3_departments = set(departements_c3[0])
filtered_data = merged_data[merged_data['LIB_SERVICE'].isin(c3_departments)]

# Assurer que la colonne LIB_SERVICE est toujours remplie
filtered_data['LIB_SERVICE'] = filtered_data['LIB_SERVICE'].fillna(method='bfill')

# Sélectionner les colonnes nécessaires pour le fichier final
final_data = filtered_data[['IGG', 'GROUP_MAIL', 'LIB_SERVICE']]

# Sauvegarder le fichier final
final_data.to_excel('C3_accredited_users.xlsx', index=False)

print("Le fichier 'C3_accredited_users.xlsx' a été créé avec succès. Le processus est terminé.")
