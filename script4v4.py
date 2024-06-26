import pandas as pd
from tqdm import tqdm
import time
from openpyxl import load_workbook

def estimate_reading_time(file_path, num_samples=100):
    """Estime le temps moyen nécessaire pour lire une ligne dans un fichier Excel."""
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active
    total_rows = ws.max_row
    
    start_time = time.time()
    for row in ws.iter_rows(min_row=2, max_row=min(num_samples+1, total_rows)):
        pass
    end_time = time.time()
    wb.close()
    
    avg_time_per_line = (end_time - start_time) / num_samples
    return avg_time_per_line, total_rows

def simulate_read_excel_with_progress(file_path, header='infer'):
    avg_time_per_line, total_rows = estimate_reading_time(file_path)
    total_time_estimate = avg_time_per_line * total_rows
    
    # Afficher la barre de progression simulée pendant le chargement
    pbar = tqdm(total=total_rows, desc=f"Lecture de {file_path}")
    data = pd.read_excel(file_path, header=header)
    for _ in range(total_rows):
        pbar.update(1)
        time.sleep(avg_time_per_line / 100)  # Diviser le temps de sommeil pour ne pas ralentir le processus
    pbar.close()

    return data

print("Initialisation du script...")

# Lire les fichiers Excel avec barres de progression
print("\n---------------\nLecture des fichiers Excel...\n---------------")
people = simulate_read_excel_with_progress('people.xlsx', header=0)
custom = simulate_read_excel_with_progress('custom.xlsx', header=0)
departements_c3 = simulate_read_excel_with_progress('departements.xlsx', header=None)

# Vérification des colonnes des DataFrames
print("\n---------------\nVérification des colonnes des DataFrames...\n---------------")
print("Colonnes de 'people':", people.columns.tolist())
print("Colonnes de 'custom':", custom.columns.tolist())

# Créer des copies des DataFrames pour les manipulations
print("\n---------------\nCréation de copies des DataFrames pour manipulation...\n---------------")
people_copy = people.copy()
custom_copy = custom.copy()

# Renommer les colonnes dans 'custom_copy'
print("\n---------------\nRenommage des colonnes dans 'custom' pour correspondre à celles de 'people'...\n---------------")
if 'GGI' in custom_copy.columns and 'Email' in custom_copy.columns and 'Department' in custom_copy.columns:
    custom_copy.rename(columns={'GGI': 'IGG', 'Email': 'GROUP_MAIL', 'Department': 'LIB_SERVICE'}, inplace=True)
    print("Colonnes après renommage dans 'custom':", custom_copy.columns.tolist())
else:
    raise KeyError("Les colonnes attendues 'GGI', 'Email' et 'Department' ne sont pas présentes dans 'custom'.")

# Fusionner people_copy avec custom_copy pour compléter les informations manquantes
print("\n---------------\nFusion des DataFrames...\n---------------")
merged_data = pd.merge(people_copy, custom_copy[['IGG', 'GROUP_MAIL', 'LIB_SERVICE']], on='IGG', how='left', suffixes=('', '_custom'))

# Utiliser fillna pour combler les informations manquantes
print("\n---------------\nComblement des informations manquantes...\n---------------")
merged_data['GROUP_MAIL'] = merged_data['GROUP_MAIL'].fillna(merged_data['GROUP_MAIL_custom'])
merged_data['LIB_SERVICE'] = merged_data['LIB_SERVICE'].fillna(merged_data['LIB_SERVICE_custom'])

# Supprimer les colonnes inutiles après la fusion
print("\n---------------\nSuppression des colonnes temporaires après la fusion...\n---------------")
merged_data.drop(columns=['GROUP_MAIL_custom', 'LIB_SERVICE_custom'], inplace=True)

# Filtrer les départements C3
print("\n---------------\nFiltration des utilisateurs selon les départements C3...\n---------------")
c3_departments = set(departements_c3.iloc[:, 0])
filtered_data = merged_data[merged_data['LIB_SERVICE'].isin(c3_departments)].copy()

# Assurer que la colonne LIB_SERVICE est toujours remplie
print("\n---------------\nAssurer que la colonne 'LIB_SERVICE' est toujours remplie...\n---------------")
filtered_data.loc[:, 'LIB_SERVICE'] = filtered_data['LIB_SERVICE'].ffill()

# Sélectionner les colonnes nécessaires pour le fichier final
print("\n---------------\nSélection des colonnes finales pour l'export...\n---------------")
final_data = filtered_data[['IGG', 'GROUP_MAIL', 'LIB_SERVICE']]

# Sauvegarder le fichier final
print("\n---------------\nSauvegarde du fichier final 'C3_accredited_users.xlsx'...\n---------------")
final_data.to_excel('C3_accredited_users.xlsx', index=False)

print("\n---------------\nLe fichier 'C3_accredited_users.xlsx' a été créé avec succès. Le processus est terminé.\n---------------")
